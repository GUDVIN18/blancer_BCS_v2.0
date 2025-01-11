from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from apps.bot_app.models import GenerationProcess, TelegramBotConfig, BotUser, Images, PromptModelSettings, LoggingProccess
from apps.stickers.models import Generate_Stickers, StikerPackConfig, Stiker_target_photo, Stiker_output_photo
from apps.bot_app.task_end_handlers import Task_Handler
from django.http import HttpResponse, JsonResponse
from django.core.files.base import ContentFile
import uuid
import re
import math
import logging
from django.db.models import Q
import random
from datetime import datetime, timedelta
from django.utils.timezone import now


from django.http import FileResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.conf import settings
import os
from django.core.files import File

from django.http import FileResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.shortcuts import get_object_or_404
from generator_v2 import leonardo_generations

#Для определения сколько на фото лиц



@csrf_exempt 
def get_task_status(request):
    if request.method == "POST":
        try:
            task_id = int(request.POST.get("id"))
            task_status = request.POST.get("task_status")
        except:
            return HttpResponse("Данные не переданы")
        
        try:
            task = GenerationProcess.objects.get(id=task_id)
            task.process_status = task_status
            task.save()

        except GenerationProcess.DoesNotExist:
            # Обработка ситуации, когда объект не найден
            print(f"GenerationProcess с id={task_id} не существует")

        print(f'________________________{task_status} {task_id}________________________\n\n')
        return HttpResponse("Success")
    else:
        return HttpResponse("Not POST")










from PIL import Image

def add_logo_to_image(file, logo_path):
    """
    Add a logo to the top right corner of an image
    
    :param file: Input image file (from request.FILES or similar)
    :param logo_path: Path to the logo image file
    :return: Modified image file
    """
    # Open the original image
    original_image = Image.open(file)
    
    # Open the logo image
    logo = Image.open(logo_path)
    
    # Resize logo (optional - adjust size as needed)
    logo_width = int(original_image.width * 0.2)  # Logo width as 20% of image width
    logo_aspect = logo.width / logo.height
    logo_height = int(logo_width / logo_aspect)
    logo = logo.resize((logo_width, logo_height), Image.LANCZOS)
    
    # Calculate position (top right corner with a small margin)
    margin = 10
    position = (
        original_image.width - logo.width - margin, 
        margin
    )
    
    # Create a copy of the original image to avoid modifying the original
    result_image = original_image.copy()
    
    # Paste the logo with transparency
    result_image.paste(logo, position, logo)
    
    # Save the result to a BytesIO object to mimic file-like object
    from io import BytesIO
    output = BytesIO()
    result_image.save(output, format='PNG')
    output.seek(0)
    
    # You might need to modify this part depending on your exact use case
    output.name = file.name  # Preserve original filename
    
    return output





import time
@csrf_exempt 
def finish_task_status(request):
    if request.method == "POST":
        print("-----ВХОД ВЫПОЛНЕН---------")
        try:
            task_id = int(request.POST.get("id"))
            task_status = request.POST.get("task_status")

            print('task_status', task_status)
        except:
            return HttpResponse("Error data")
        
        try:
            wait_interval = 5  # интервал проверки (1 секунда)
            task = GenerationProcess.objects.get(id=task_id)

            task.process_status = task_status
            task.save()

        except GenerationProcess.DoesNotExist:
            # Обработка ситуации, когда объект не найден
            print(f"GenerationProcess с id={task_id} не существует")
        
        print('task_status', task_status)

        task_end_handler = task.task_end_handler
        if task_status == "ERROR_GENERATION":
            task_end_handler = 'error_generations'
            handler = getattr(Task_Handler(), task_end_handler)
            handler(task=task)
            return HttpResponse('OK')

        try:
            print('\n\n\nrequest.FILES', request.FILES, '\n\n\n\n')
            file = request.FILES.get("file")
            try:
                print(f'\n\n123file {file}')
                output_image = Images.objects.create(
                    description='Фото с логотипом',
                    image=file
                )

                # Update task if needed
                task.output_photo = output_image
                task.save()
            except Exception as e:
                return HttpResponse(f"Данные не переданы, {e}")
            
            print(f'________________________{task_status} {task_id}________________________\n\n')
            

            if task_end_handler:
                handler = getattr(Task_Handler(), task_end_handler)
                handler(task=task)

            return HttpResponse('OK')
        
        except Exception as e:
            if task_end_handler:
                
                handler = getattr(Task_Handler(), task_end_handler)
                handler(task=task)
            print('Файл не был передан', e)
            return HttpResponse('Not OK')



            

    else:
        return HttpResponse("Not POST")
    


# def count_faces(image):
#     # Если это объект изображения, преобразуем его в массив
#     if isinstance(image, Image.Image):
#         img_array = np.array(image)
#     else:
#         img_array = face_recognition.load_image_file(image)  # Загружаем изображение по пути
    
#     # Найдем все лица
#     face_locations = face_recognition.face_locations(img_array)
    
#     # Вернем количество найденных лиц



import openpyxl








@csrf_exempt 
def create_task(request):
    if request.method == "POST":
        try:
            # target_photo_file = request.FILES.get("target_photo") #Фото обезьяны
            prompt = request.POST.get("prompt")
            user_id = request.POST.get("user_id")
            format_photo = request.POST.get("format_photo")
            task_end_handler = request.POST.get("task_end_handler")
            # prompt_number = request.POST.get("promt_num")
            user_photo_file = request.FILES.get("user_photo") #Фото пользователя

            #Если True - генерация , если False - замена лица на готовом фото
            generation_or_face_to_face = request.POST.get("generation_or_face_to_face")
            print('generation_or_face_to_face', generation_or_face_to_face, type(generation_or_face_to_face))

            print('user_prompt', prompt)

            try:
                if prompt.split(' ')[0] == 'promt_user':
                    data_part = prompt.split(' ')[1]  # "1_20_Эмоции_Консервативный_Мужчина_Горизонтальный"
                    parts = data_part.split('_')      # ["1", "20", "Эмоции", "Консервативный", "Мужчина", "Горизонтальный"]

                    summa = parts[0]  # "1"
                    period = parts[1]  # "20"
                    interes = parts[2]  # "Эмоции"
                    type_investor = parts[3]  # "Консервативный"
                    gender = parts[4]  # "Мужчина"
                    format_photo = parts[5]  # "Горизонтальный"

                    print('summa:', summa)
                    print('period:', period)
                    print('interes:', interes)
                    print('type_investor:', type_investor)
                    print('gender:', gender)
                    print('format_photo:', format_photo)


                    if interes == "Эмоции":
                        interes = "Эмоции / Вдохновение"


                    promt_models = PromptModelSettings.objects.filter(
                        # Интерес совпадает с interest_1 или interest_2
                        Q(interest_1=interes) | Q(interest_2=interes),
                        # Бюджет совпадает с одним из значений budget_1, budget_2 или budget_3
                        Q(budget_1=summa) | Q(budget_2=summa) | Q(budget_3=summa)
                    )

                    random_model = random.choice(promt_models)

                    if random_model.number == 67:
                        random_model = random.choice(promt_models)

                    base_dir = '/root/project/balancer-v2.0/face_to_face_server0/media'
                    prompt_folder = os.path.join(base_dir, str(random_model.number))
                    # prompt_folder = os.path.join(base_dir, str(prompt_number))


    
                    # if gender == "Мужчина":
                    #     prompt = random_model.men_promt
                    #     print('\n\n\n', prompt)
                    # elif gender == "Женщина":
                    #     prompt = random_model.women_promt


                    if gender == "Мужчина":
                        gender_folder = os.path.join(prompt_folder, 'men')
                    elif gender == "Женщина":
                        gender_folder = os.path.join(prompt_folder, 'women')


                    gender_folder_format_photo = os.path.join(gender_folder, format_photo)


                    available_folders = [
                        folder for folder in os.listdir(gender_folder_format_photo)
                        if os.path.isdir(os.path.join(gender_folder_format_photo, folder)) and not folder.endswith('.zip')
                    ]
                    print('available_folders:', available_folders)

                    # Если есть доступные папки
                    if available_folders:
                        # Выбираем случайную папку из списка
                        random_folder = random.choice(available_folders)
                        
                        # Формируем путь к случайной папке
                        random_folder_path = os.path.join(gender_folder_format_photo, random_folder)
                        print('random_folder_path:', random_folder_path)

                    else:
                        random_folder_path = gender_folder_format_photo
                        
                    try:
                        all_files = os.listdir(random_folder_path)

                        try:
                            last_proccess_user = GenerationProcess.objects.filter(user_id=user_id).order_by('-id').first()
                            last_photo_path = last_proccess_user.target_photo.image.path
                        except Exception as e:
                            print('У пользователя не было полседней записи')
                            last_photo_path = None
                        
                        
                        # Отбираем файлы, в названии которых есть нужный формат
                        matching_files = [f for f in all_files]
                        if matching_files:
                            selected_file = random.choice(matching_files)
                            full_file_path_dont_user = os.path.join(random_folder_path, selected_file)

                            print(f'\n\n\n ВЫБРАННЫЙ НОМЕР ПРОМТА {random_model.number}\nПОЛНЫЙ ПУТЬ {random_folder_path}\nВСЕ ФАЙЛЫ {matching_files}\nРАНДОМНЫЙ ФАЙЛ {selected_file}\nПОЛНЫЙ ПУТЬ ДО ФАЙЛА {full_file_path_dont_user}\n\n\n')
                            if last_photo_path != full_file_path_dont_user or last_photo_path == None:
                                full_file_path = full_file_path_dont_user
                            else:
                                selected_file_2 = random.choice(matching_files)
                                full_file_path_dont_user_2 = os.path.join(random_folder_path, selected_file_2)
                                full_file_path = full_file_path_dont_user_2
                                full_file_path_dont_user_2  # Use the first one as fallback

                        
                            print(f'Выбранный файл full_file_path: {full_file_path}')
                        else:
                            print(f'Не найдено ни одного файла с форматом {format_photo} в папке {gender_folder_format_photo}')
                    except Exception as e:
                        print('Ошибка с формированием пути', e)

                    
                    textovka = random_model.text
                    
                    print('gender promt')
                


                elif prompt.split(' ')[0] == 'generation_10_photo':
                    user_id = request.POST.get("user_id")
                    user_photo = request.FILES.get("user_photo") #Фото пользователя
                    photo_target = request.FILES.get("photo") #Фото на кого надо

                    user_photo = Images.objects.create(
                        description = 'Фото пользователя для 10 фоток',
                        image=user_photo,
                    )

                    target_photo = Images.objects.create(
                        description = 'Фото на кого будет наложено для 10 фоток',
                        image=photo_target,
                    )
                    textovka = "стать ... Тест через скрипт 10 фото! Без текстовки"
                    task_end_handler = 'task_end_alert'

                    new_generation = GenerationProcess(
                        textovka_new=textovka,
                        user_id = user_id,
                        photo = user_photo,
                        target_photo = target_photo,
                        format_photo = 'Горизонтальный',
                        process_status='WAITING',
                        process_backend_id=uuid.uuid4(),
                        task_end_handler = task_end_handler,
                    )
                    
                    new_generation.save()
                    return JsonResponse({"success": new_generation.id}, status=200)





                else:
                    textovka = "стать ... Тест через скрипт! Без текстовки"
                    summa = '1'
                    period = "20"
                    interes = "Эмоции / Вдохновение"
                    type_investor = 'Консервативный'
                    gender = 'Мужчина'
                    full_file_path = '/root/project/balancer-v2.0/face_to_face_server0/media/saha_tai.jpg'

            except Exception as e:
                print(f'Ошибка при выборке промта, {e}')
            
                

        except (ValueError, TypeError):
            return JsonResponse({"error": "Неверный формат данных"}, status=400)
        

        
        try:
            #Получаем, сколько пользователей перед пользователем
            user_waiting = GenerationProcess.objects.filter(process_status='WAITING').count()
            print(f'\n ---- ПОЛЬЗОВАТЕЛЕЙ В ОЧЕРЕДИ ПЕРЕД НОВОЙ ГЕНЕРЦИЕЙ {user_waiting}\n')

            
            user_photo = Images.objects.create(
                description = 'Фото пользователя',
                image=user_photo_file,
            )

            if type_investor == "Умеренный":
                i = 0.3/12
            if type_investor == "Рискованный":
                i = 0.4/12
            if type_investor == "Консервативный":
                i = 0.2/12

            mapping = {
                '25': 25000,
                '50': 50000,
                '100': 100000,
                '250': 250000,
                '500': 500000,
                '1': 1000000,
            }

            summa_str = summa  # 'summa' уже содержится в переменной из логики
            if summa_str in mapping:
                c = float(mapping[summa_str])

            n = int(period) * 12
            print(c, type(c))
            fV = round(c * (((1 + i)**n - 1) / i) * (1 + i))
            fv = f"{fV:,.0f}".replace(",", " ")

            print('ФОРМУЛА', fv)

            print('random_textovka', textovka)
            try:
                if textovka.split(' ')[0] == 'на':
                    textovka_new = f'У Вас будет {fv}₽, этой суммы хватит {textovka}'

                elif textovka.split(' ')[0] == 'стать':
                    textovka_new = f'У Вас будет {fv}₽, этой суммы хватит чтобы {textovka}'
                
                else:
                    textovka_new = f'У Вас будет {fv}₽, этой суммы хватит {textovka}'
            except Exception as e:
                print('Ошибка во views textovka_new', e)


            if generation_or_face_to_face == 'True':
                negative_prompt = request.POST.get("negative_prompt")

                new_generation = GenerationProcess(
                    # target_photo=target_photo,
                    prompt = prompt,
                    negative_prompt=negative_prompt,
                    textovka_new=textovka_new,
                    user_id = user_id,
                    photo = user_photo,
                    format_photo = format_photo,
                    process_status='WAITING',
                    process_backend_id=uuid.uuid4(),
                    task_end_handler = task_end_handler,
                )
                
                new_generation.save()



            # target_path = '/root/project/balancer-v2.0/face_to_face_server0/media/saha_tai.jpg'

            with open(full_file_path, 'rb') as f:
                target_photo = Images.objects.create(
                    description='Фото на кого будет наложено',
                    image=File(f, name='full_file_path.jpg')
                )

                if generation_or_face_to_face == 'False':

                    new_generation = GenerationProcess(
                        # target_photo=target_photo,
                        textovka_new=textovka_new,
                        user_id = user_id,
                        photo = user_photo,
                        target_photo = target_photo,
                        format_photo = format_photo,
                        process_status='WAITING',
                        process_backend_id=uuid.uuid4(),
                        task_end_handler = task_end_handler,
                    )
                    
                    new_generation.save()



            try:
                loger = LoggingProccess(
                    user_id=user_id,
                    generation_number = new_generation,
                    result_formula = fv,
                    user_price = str(c),
                    user_category = interes,
                    gender = gender,
                    time_invest = period,
                    investor_risk = type_investor,
                    textovka_new = textovka_new,
                )
                loger.save()
            except:
                loger = LoggingProccess(
                    user_id=user_id,
                    generation_number = new_generation,
                    user_price = str(c),
                    user_category = interes,
                    gender = gender,
                    time_invest = period,
                    investor_risk = type_investor,
                    textovka_new = textovka_new,
                )
                loger.save()


        
        except Exception as e:
            print(f"Ошибка 349 {e}")
            return JsonResponse({"error": str(e)}, status=500)
        
        print(f'________________________{new_generation.process_status} {new_generation.id}________________________\n\n')

        return JsonResponse({"task_id": new_generation.id, "status": new_generation.process_status, "user_waiting": user_waiting})
    else:
        return JsonResponse({"error": "Метод не разрешен"}, status=405)



from django.utils import timezone
from datetime import datetime
@csrf_exempt
@require_http_methods(["GET"])
def user_waiting(request):
    # Максимально допустимое количество процессов на сервер
    MAX_PROCESSES = 11

    # Принятых генераций всего
    procces_accepted = GenerationProcess.objects.filter(process_status='ACCEPTED').count()
    # Завершенных генераций
    proccess_completed = GenerationProcess.objects.filter(process_status='COMPLETED').count()
    # Генераций с ошибкой ERROR_GENERATION
    proccess_error = GenerationProcess.objects.filter(process_status='ERROR_GENERATION').count()
    # Очередь
    user_waiting = GenerationProcess.objects.filter(process_status='WAITING').count()

    # Получаем текущее время 
    current_time = timezone.now()

    # Генерации за сегодня
    procces_accepted_today = GenerationProcess.objects.filter(
        process_status='COMPLETED',
        process_start_time__date=current_time.date()
    ).count()

    # Генерации за вчера 
    procces_accepted_yesterday = GenerationProcess.objects.filter(
        process_status='COMPLETED',
        process_start_time__date=current_time.date() - timedelta(days=1)
    ).count()

    # Серверы и их загрузка
    servers = {
        "server_1": GenerationProcess.objects.filter(process_status='ACCEPTED', server_int=11).count(),
        "server_2": GenerationProcess.objects.filter(process_status='ACCEPTED', server_int=10).count(),
        "server_4": GenerationProcess.objects.filter(process_status='ACCEPTED', server_int=8).count(),
        "server_5": GenerationProcess.objects.filter(process_status='ACCEPTED', server_int=7).count(),
        "server_6": GenerationProcess.objects.filter(process_status='ACCEPTED', server_int=2).count(),
        "server_7": GenerationProcess.objects.filter(process_status='ACCEPTED', server_int=12).count(),
        "server_8": GenerationProcess.objects.filter(process_status='ACCEPTED', server_int=15).count(),
        "server_9": GenerationProcess.objects.filter(process_status='ACCEPTED', server_int=16).count(),
        "server_10": GenerationProcess.objects.filter(process_status='ACCEPTED', server_int=17).count(),
    }

    # Анализ загрузки серверов
    server_percentages = {
        server_name: round(min((count / MAX_PROCESSES) * 100, 100), 2)  # Ограничение 100%
        for server_name, count in servers.items()
    }
    
    data = {
        "user_waiting": user_waiting,
        "procces_accepted": procces_accepted,
        "proccess_completed": proccess_completed,
        "proccess_error": proccess_error,
        "procces_accepted_yesterday": procces_accepted_yesterday,
        "procces_accepted_today": procces_accepted_today,
        **server_percentages,  # Добавляем данные серверов с процентами
    }

    return JsonResponse(data)



from generator import main

@csrf_exempt
def start_generator(request):
    # promt_number = request.GET.get('promt_number')
    # photo_count = request.GET.get('photo_count')
    # now_num = request.GET.get('now_num')
    # men = request.GET.get('men') 
    # women = request.GET.get('women') 
    # main(promt_number, photo_count, now_num, men, women)
    Images.objects.filter(id__lte=178576).delete()
    GenerationProcess.objects.filter(user_id=-1002423543289).delete()
    return HttpResponse("start_generator OK", status=200)





@csrf_exempt
@require_http_methods(["GET"])
def get_task_result(request):
    print('Зашли в get_task_result')
    # Получаем идентификатор задачи из запроса
    task_id = request.GET.get('task_id')
    print(task_id)
    if not task_id:
        print('Не указан идентификатор задачи')
        return HttpResponse("Не указан идентификатор задачи", status=400)
    
    # Получаем объект Task или возвращаем 404, если не найден
    try:
        generation_object = GenerationProcess.objects.get(
            id=int(task_id),
        )
    except Exception as e:
        print('GenerationProcess не найден')
        return HttpResponse("GenerationProcess не найден", status=404)
    
    # Проверяем, есть ли изображение в задаче
    if not generation_object.output_photo:
        print('Изображение не найдено для данной задачи')
        return HttpResponse("Изображение не найдено для данной задачи", status=404)
    
    # Открываем файл изображения
    try:
        file = generation_object.output_photo.image.open()
    except IOError:
        print('Ошибка при открытии файла')
        return HttpResponse("Ошибка при открытии файла", status=500)
    
    # Определяем тип содержимого на основе расширения файла
    file_name = generation_object.output_photo.image.path
    content_type = 'image/jpeg' if file_name.lower().endswith(('.jpg', '.jpeg')) else 'image/png'
    
    # Создаем FileResponse
    response = FileResponse(file)
    response['Content-Type'] = content_type
    response['Content-Disposition'] = f'attachment; filename="{file_name}"'

    print('ВСЕ ОК')
    
    return response





import requests


@csrf_exempt
def create_task_multidata(request):
    if request.method == "POST":
        print('Мы в функции')

        try:
            bot_token = 'bot8049871685:AAHwm0S0CICmXFlGsKTgHxJ5JokyCT6pkSM'
            print("Form data:", request.POST, '\n')

            data = request.POST
            form_value = list(data.keys())[-1]  # Берем последний ключ

            user_id = form_value.split("\n")[0] 
            price = form_value.split("\n")[1] 
            time_invest = form_value.split("\n")[2] 
            category_invest = form_value.split("\n")[3]
            investor = form_value.split("\n")[4]
            gender = form_value.split("\n")[5]
            file_id = form_value.split("\n")[6]
            index_amount = form_value.split("\n")[7]
            index_interest = form_value.split("\n")[8]
            prompt = form_value.split("\n")[-1]

            number = re.search(r'\d+', time_invest)
            if number:
                number = int(number.group()) 
            else:
                number = int(0)

            print(f"User ID: {user_id}")
            print(f"Price: {price}")
            print(f"Time Invest: {number}")
            print(f"Category Invest: {category_invest}")
            print(f"Investor: {investor}")
            print(f"Gender: {gender}")
            print(f"File ID: {file_id}")
            print(f"Index Amount: {index_amount}")
            print(f"Index Interest: {index_interest}")
            print(f"Prompt: {prompt}")

            
            # Define mappings for index_amount to prices
            amount_mapping = {
                "Z": 25000,
                "X": 50000,
                "C": 100000,
                "V": 250000,
                "F": 500000,
                "L": 1000000
            }

            # Define mappings for index_interest to categories
            interest_mapping = {
                "A": "Путешествия",
                "B": "Недвижимость",
                "C": "Развлечения",
                "D": "Автомобили",
                "E": "Саморазвитие",
                "F": "Эмоции/Вдохновение"
            }


            # Get the corresponding price and category
            c = amount_mapping.get(index_amount, "Unknown amount")
            user_category = interest_mapping.get(index_interest, "Unknown interest")

            # Print the results

            print(f"C: {c}")

            if investor == "Умеренный":
                i = 0.3
            elif investor == "Рискованный":
                i = 0.2
            elif investor == "Консервативный":
                i = 0.4
            
            n = number*12



            # Шаги расчета
            # 1. Вычисляем (1 + i)
            one_plus_i = 1 + i

            # 2. Возводим (1 + i) в степень n
            power_term = one_plus_i ** n

            # 3. Вычисляем ((1 + i)^n - 1)
            numerator = power_term - 1

            # 4. Делим результат на i
            fraction = numerator / i

            # 5. Умножаем результат на C
            main_term = c * fraction

            # 6. Умножаем на (1 + i)
            FV_Annuity_Due = main_term * one_plus_i

            print('INFO FORMULA')
            print('i', i)
            print('n', n)
            print(f"C: {c}")


            print('RESULT_FORMULA', FV_Annuity_Due)









            filtered_objects = PromptModelSettings.objects.filter(
                budget=index_amount,  # Фильтруем по бюджету
                interest=index_interest  # Фильтруем по интересу
            )




            get_file_url = f'https://api.telegram.org/{bot_token}/getFile?file_id={file_id}'
            response = requests.get(get_file_url)
            file_info = response.json()
            print(f'\n\nfile_info {file_info}')

            if file_info["ok"]:
                file_path = file_info["result"]["file_path"]
                file_url = f'https://api.telegram.org/file/{bot_token}/{file_path}'
                file_response = requests.get(file_url)

                if file_response.status_code == 200:
                    print(settings.MEDIA_ROOT)
                    file_path = f'{settings.MEDIA_ROOT}/img_storage/{user_id}.jpg'
                    with open(file_path, 'wb') as f:
                        f.write(file_response.content)
                    print("Файл успешно скачан!")
                else:
                    print("Ошибка при скачивании файла")
            else:
                print("Ошибка получения информации о файле")


            task_end_handler = "task_end_alert"
            user_photo_file = f'img_storage/{user_id}.jpg'
            

            try:
                #Получаем, сколько пользователей перед пользователем
                user_waiting = GenerationProcess.objects.filter(process_status='WAITING').count()
                print(f'\n ---- ПОЛЬЗОВАТЕЛЕЙ В ОЧЕРЕДИ ПЕРЕД НОВОЙ ГЕНЕРЦИЕЙ {user_waiting}\n')
                
                user_photo = Images.objects.create(
                    description = 'Фото пользователя',
                    image=user_photo_file,
                )

                new_generation = GenerationProcess(
                    # target_photo=target_photo,
                    prompt = prompt,
                    photo = user_photo,
                    process_status='WAITING',
                    process_backend_id=uuid.uuid4(),
                    task_end_handler = task_end_handler,
                    user_id = user_id,
                )
                
                new_generation.save()



                LoggingProccess.objects.create(
                user_id=user_id,
                generation_number = new_generation,
                generation_time = None,
                result_formula = FV_Annuity_Due,
                user_price = c,
                user_category = user_category,
                gender = gender,
                time_invest = time_invest,
                investor_risk = investor,

                )

                print(f'________________________{new_generation.process_status} {new_generation.id}________________________\n\n')

                return JsonResponse({"task_id": new_generation.id, "status": new_generation.process_status, "user_waiting": user_waiting})

                
            
            except Exception as e:
                return JsonResponse({"error": str(e)}, status=500)

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    else:
        return JsonResponse({"error": "method no aloved"}, status=405)


import urllib.parse
@csrf_exempt
def deletepath(request, path):
    try:
        # URL-декодирование
        decoded_path = urllib.parse.unquote(path)
        
        # Добавляем "/" в начало пути, если его нет
        if not decoded_path.startswith('/'):
            decoded_path = '/' + decoded_path
        
        # Используем абсолютный путь с корректным разделением
        full_path = os.path.abspath(decoded_path)
        
        # Проверяем существование файла перед удалением
        if os.path.exists(full_path):
            os.remove(full_path)
            return JsonResponse({"status": "OK"})
        else:
            return JsonResponse({"status": "error", "message": f"File not found: {full_path}"}, status=404)
    
    except PermissionError:
        return JsonResponse({"status": "error", "message": "Permission denied"}, status=403)
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=500)


# from django.http import JsonResponse
# from django.views.decorators.csrf import csrf_exempt

# @csrf_exempt
# def create_task_multidata(request):
#     print("Вызвана функция create_task_multidata")
#     if request.method == "POST":
#         print("Это POST-запрос")
#         if request.content_type == "multipart/form-data":
#             print("Тип контента multipart/form-data")
#             return JsonResponse({"message": "Успешно обработано"}, status=200)
#         else:
#             print(f"Тип контента: {request.content_type}")
#             return JsonResponse({"error": "Неправильный тип контента"}, status=400)
#     return JsonResponse({"error": "Метод не поддерживается"}, status=405)














@csrf_exempt
def task_status(request, task_id):
    """
    Получение статуса задачи по её ID
    """
    try:
        task = GenerationProcess.objects.get(id=task_id)
        
        return JsonResponse({
            'status': task.process_status,
            'task_id': task.id
        })
    
    except GenerationProcess.DoesNotExist:
        return JsonResponse({
            'status': 'ERROR',
            'message': 'Задача не найдена'
        }, status=404)
    



# @csrf_exempt 
# def get_file(request, task_id):
#     """
#     Получение файла сгенерированного изображения по ID задачи
#     """
#     try:
#         task = GenerationProcess.objects.get(id=task_id)
        
#         if not task.output_photo:
#             return JsonResponse({
#                 'status': 'ERROR', 
#                 'message': 'Файл не найден'
#             }, status=404)
        
#         with open(task.output_photo.path, 'rb') as file:
#             response = HttpResponse(
#                 file.read(), 
#                 content_type='image/png'
#             )
#             response['Content-Disposition'] = f'attachment; filename="{task.output_photo.name}"'
#             return response
    
#     except GenerationProcess.DoesNotExist:
#         return JsonResponse({
#             'status': 'ERROR', 
#             'message': 'Задача не найдена'
#         }, status=404)
#     except FileNotFoundError:
#         return JsonResponse({
#             'status': 'ERROR', 
#             'message': 'Физический файл отсутствует'
#         }, status=404)



from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse, JsonResponse
from django.core.files.storage import default_storage
from PIL import Image
import io


from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
import os

@csrf_exempt
def get_file(request, task_id):
    print('task_id',task_id)
    try:
        
        task = GenerationProcess.objects.get(id=task_id)
        
        if not task.output_photo:
            return HttpResponse('Файл не найден', status=404)
        
        file_path = task.output_photo.image.path
        
        if not file_path or not os.path.exists(file_path):
            return HttpResponse('Физический файл отсутствует', status=404)

        # Просто отдаем файл без повторной валидации
        with open(file_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='image/png')
            response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
            return response
    
    except GenerationProcess.DoesNotExist:
        return HttpResponse('Задача не найдена', status=404)
    except Exception as e:
        print(f"Непредвиденная ошибка: {e}")
        return HttpResponse('Внутренняя ошибка сервера', status=500)
    









from datetime import datetime

@csrf_exempt
def get_logs(request):
    if request.method != "POST":
        return HttpResponse('Требуется POST запрос', status=400)
    
    file_path = '/root/project/balancer-v2.0/face_to_face_server0/xlsx/log.xlsx'
    
    try:
        # Получаем данные
        logs = LoggingProccess.objects.all()
        
        # Создаем Excel файл
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Заголовки
        headers = [
            'Telegram ID',
            'Процесс',
            'Старт',
            'Время выполнения',
            'Результат формулы',
            'Сумма',
            'Интерес',
            'Гендер',
            'На какой срок',
            'Риски',
            'Готовая текстовка'
        ]
        worksheet.append(headers)

        # Записываем данные
        for log in logs:
            row = [
                str(log.user_id) if log.user_id else '',
                str(log.generation_number) if log.generation_number else '',
                log.process_start_time.replace(tzinfo=None) if log.process_start_time else '',
                str(log.generation_time) if log.generation_time else '',
                str(log.result_formula) if log.result_formula else '',
                str(log.user_price) if log.user_price else '',
                str(log.user_category) if log.user_category else '',
                str(log.gender) if log.gender else '',
                str(log.time_invest) if log.time_invest else '',
                str(log.investor_risk) if log.investor_risk else '',
                str(log.textovka_new) if log.textovka_new else ''
            ]
            worksheet.append(row)

        # Настраиваем ширину колонок
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Создаем директорию если её нет
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        # Сохраняем файл
        workbook.save(file_path)
        
        # Отправляем файл
        response = FileResponse(
            open(file_path, 'rb'),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="log.xlsx"'
        return response

    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Ошибка: {str(e)}")
        print(f"Детали:\n{error_details}")
        return HttpResponse(f'Ошибка сервера: {str(e)}', status=500)

    finally:
        # Удаляем временный файл если он существует
        if os.path.exists(file_path):
            try:
                os.remove(file_path)
            except:
                pass




import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse

def export_prompt_model_settings_xlsx(request):
    """
    Выгружает данные модели PromptModelSettings в формате XLSX.
    """
    # Создаём новую книгу Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Prompt Model Settings"

    # Добавляем заголовки
    headers = [
        'ID', 'Номер промта', 'Ролевая', 'Назначение',
        'Мужской Промпт', 'Женский Промпт', 
        'Интерес 1', 'Интерес 2', 
        'Бюджет 1', 'Бюджет 2', 'Бюджет 3', 'После'
    ]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)  # Делаем заголовки жирными

    # Заполняем данными
    for row_num, setting in enumerate(PromptModelSettings.objects.all(), 2):
        sheet.cell(row=row_num, column=1, value=setting.id)
        sheet.cell(row=row_num, column=2, value=setting.number)
        sheet.cell(row=row_num, column=3, value=setting.rolevaya)
        sheet.cell(row=row_num, column=4, value=setting.purpose)
        sheet.cell(row=row_num, column=5, value=setting.men_promt)
        sheet.cell(row=row_num, column=6, value=setting.women_promt)
        sheet.cell(row=row_num, column=7, value=setting.interest_1)
        sheet.cell(row=row_num, column=8, value=setting.interest_2)
        sheet.cell(row=row_num, column=9, value=setting.budget_1)
        sheet.cell(row=row_num, column=10, value=setting.budget_2)
        sheet.cell(row=row_num, column=11, value=setting.budget_3)
        sheet.cell(row=row_num, column=12, value=setting.text)

    # Создаём HTTP-ответ
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="prompt_model_settings.xlsx"'

    # Сохраняем книгу в ответ
    workbook.save(response)

    return response







@csrf_exempt
def start_leonardo_generations(request):
    try:
        if request.method == "POST":
            print('Зашли в start_leonardo_generations')
            
            # Получаем параметры из GET запроса
            user_id = request.POST.get('user_id', '')  # Добавляем значение по умолчанию
            prompt = request.POST.get('prompt', '')
            negative_prompt = request.POST.get('negative_prompt', '')
            orientation = request.POST.get('format_photo', '')
            
            print(f"Полученные параметры:")
            print(f"user_id: {user_id}")
            print(f"prompt: {prompt}")
            print(f"negative_prompt: {negative_prompt}")
            print(f"orientation: {orientation}")

            # Проверяем обязательные параметры
            if not all([user_id, prompt, orientation]):
                return HttpResponse("Missing required parameters", status=400)

            file_path = None
            if 'user_photo' in request.FILES:
                uploaded_file = request.FILES['user_photo']
                file_directory = f"/root/project/balancer-v2.0/face_to_face_server0/media/leonardo_kino_xl/user_photo/"
                os.makedirs(file_directory, exist_ok=True)
                file_path = os.path.join(file_directory, f'user_photo_leonardo_{user_id}.png')
                
                with open(file_path, 'wb+') as destination:
                    for chunk in uploaded_file.chunks():
                        destination.write(chunk)
                print(f"Файл сохранен: {file_path}")

            number_photo = 4
            gender = user_id
            num_gen = random.randint(1, 999999999999999999999)

            print("Запуск первой генерации")
            status = leonardo_generations(
                prompt=prompt,
                negative_prompt=negative_prompt,
                orientation=orientation,
                number_photo=number_photo,
                gender=gender,
                num_gen=num_gen,
                file_path=file_path
            )

            if status:
                print("Запуск второй генерации")
                num_gen = random.randint(1, 999999999999999999999)
                leonardo_generations(
                    prompt=prompt,
                    negative_prompt=negative_prompt,
                    orientation=orientation,
                    number_photo=number_photo,
                    gender=gender,
                    num_gen=num_gen,
                    file_path=file_path
                )
                return HttpResponse("Success run 8 photo on leonardo ai", status=200)
            else:
                return HttpResponse("Generation failed", status=500)

        return HttpResponse("Method not allowed", status=405)

    except Exception as e:
        import traceback
        print(f"Error: {str(e)}")
        print(f"Traceback: {traceback.format_exc()}")
        return HttpResponse(f"Server error: {str(e)}", status=500)