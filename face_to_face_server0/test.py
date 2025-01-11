"""Django's command-line utility for administrative tasks."""
import os
import sys
import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'face_to_face_server0.settings')
django.setup()

from apps.bot_app.models import *
from apps.stickers.models import Generate_Stickers, StikerPackConfig, Stiker_target_photo, Stiker_output_photo
from apps.bot_app.models import TelegramBotConfig, BotUser, GenerationProcess, Images, PromptModelSettings
import requests
import json



# def task_end_alert(task):
#     os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'face_to_face_server0.settings')
#     print('Отправка пользователю')
#     token = "CBV4mMeOSHNB2qksidGudZXyqhjufctJ"
#     caption = 'Ваша фото успешно готова!'

#     url = f'https://api.puzzlebot.top/?token={token}&method=tg.sendPhoto'

#     photo_url = f'https://balancer-bcs.ru{task.output_photo.image.url}',
#     print(photo_url)
#     data = {
#         "chat_id": task.user_id,
#         "caption": caption,
#         "photo": photo_url,
#     }
#     response = requests.post(url, data=data)
#     print(response.json())



# task = GenerationProcess.objects.get(id=22)
# task_end_alert(task)



import openpyxl

# Загрузка Excel-файла
file_path = "/root/project/balancer-v2.0/face_to_face_server0/pravky.xlsx"  # Замените на путь к вашему файлу
wb = openpyxl.load_workbook(file_path)
sheet = wb.active

# Определяем индексы столбцов для нужных данных
col_rolevaya = "C"
col_purpose = "D"  
col_text = "E"  
col_interest_1 = "G"
col_interest_2 = "H"
col_budget_1 = "I"
col_budget_2 = "J"
col_budget_3 = "K"
col_men_promt = "L"
col_women_promt = "M"

# Перебор строк и извлечение данных
output_lines = []
for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row):  # Пропускаем первые три строки (например, заголовки)
    purpose = row[openpyxl.utils.cell.column_index_from_string(col_purpose) - 1].value
    rolevaya = row[openpyxl.utils.cell.column_index_from_string(col_rolevaya) - 1].value
    text = row[openpyxl.utils.cell.column_index_from_string(col_text) - 1].value

    try:
        interest_1 = row[openpyxl.utils.cell.column_index_from_string(col_interest_1) - 1].value.capitalize()
    except:
        interest_1 = row[openpyxl.utils.cell.column_index_from_string(col_interest_1) - 1].value
    try:
        interest_2 = row[openpyxl.utils.cell.column_index_from_string(col_interest_2) - 1].value.capitalize()
    except:
        interest_2 = row[openpyxl.utils.cell.column_index_from_string(col_interest_2) - 1].value
        
    budget_1 = row[openpyxl.utils.cell.column_index_from_string(col_budget_1) - 1].value
    budget_2 = row[openpyxl.utils.cell.column_index_from_string(col_budget_2) - 1].value
    budget_3 = row[openpyxl.utils.cell.column_index_from_string(col_budget_3) - 1].value
    men_promt_value = row[openpyxl.utils.cell.column_index_from_string(col_men_promt) - 1].value
    women_promt_value = row[openpyxl.utils.cell.column_index_from_string(col_women_promt) - 1].value

    # Если поле "Интерес 1" пустое, останавливаемся
    if not interest_1:
        break

    # Формируем строку
    output = (
        f"Строка {row[0].row}\n"
        f"Назначение: {purpose}\n"
        f"Интерес 1: {interest_1}\n"
        f"Интерес 2: {interest_2}\n"
        f"Бюджет 1: {budget_1}\n"
        f"Бюджет 2: {budget_2}\n"
        f"Бюджет 3: {budget_3}\n"
        f"Текст: {text}\n"
        f"М Промпт: {men_promt_value}\n"
        f"Ж Промпт: {women_promt_value}\n"
    )

    # Вывод данных (опционально)

    try:
        budget_1 = budget_1.split()[0]
    except:
        budget_1 = budget_1


    try:
        budget_2 = budget_2.split()[0]
    except:
        budget_2 = budget_2



    try:
        budget_3 = budget_3.split()[0]
    except:
        budget_3 = budget_3


    text_zagotovka = "Эмоции/вдохновение" 
    if interest_1 == text_zagotovka:

        # Разделяем строку по "/"
        parts = text_zagotovka.split('/')

        # Приводим каждую часть к формату с заглавной буквы
        formatted_parts = [part.capitalize() for part in parts]

        # Соединяем части с пробелом и символом "/"
        interest_1 = ' / '.join(formatted_parts)


    if interest_2 == text_zagotovka:

        # Разделяем строку по "/"
        parts = text_zagotovka.split('/')

        # Приводим каждую часть к формату с заглавной буквы
        formatted_parts = [part.capitalize() for part in parts]

        # Соединяем части с пробелом и символом "/"
        interest_2 = ' / '.join(formatted_parts)


    # Создание объекта в базе данных
    PromptModelSettings.objects.create(
        number=int(row[0].row) - 3,
        rolevaya=rolevaya,
        purpose=purpose,
        men_promt=men_promt_value,
        women_promt=women_promt_value,
        interest_1 = interest_1,
        interest_2 = interest_2,
        budget_1=budget_1,
        budget_2=budget_2,
        budget_3=budget_3,
        text=text,
    )


    print(output)



# url = 'http://91.218.245.239:9999/task_error_alert'
            
# data = {
#     "chat_id": 6424595615,
# }

# # Отправка запроса

# response = requests.post(url, data=data)
# print(response.json())